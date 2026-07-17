import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import fitz
import openpyxl

class RoboSupremoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Robô Supremo - Filtro Interno e PDF Automático")
        self.root.geometry("800x600")
        
        self.pdf_molde = ""
        self.excel_dados = ""
        self.map_file = "mapeamento_supremo.xlsx"
        self.df_base = None
        
        tk.Label(root, text="Robô Supremo: O Motor Completo (Filtro Interno + PDF)", font=("Arial", 16, "bold")).pack(pady=10)
        tk.Label(root, text="Lê a Base de Dados pura, faz os cálculos da revisão e carimba no Molde.", font=("Arial", 10)).pack(pady=5)
        
        # Arquivos
        frame_arq = tk.LabelFrame(root, text=" 1. Seleção de Ficheiros ", padx=10, pady=10)
        frame_arq.pack(fill="x", padx=20, pady=5)
        
        self.lbl_pdf = tk.Label(frame_arq, text="PDF Molde: Não selecionado", fg="red")
        self.lbl_pdf.grid(row=0, column=0, sticky="w")
        tk.Button(frame_arq, text="Selecionar Lâmina PDF", command=self.load_pdf).grid(row=0, column=1, padx=10)
        
        self.lbl_excel = tk.Label(frame_arq, text="Excel (Base Gigante): Não selecionado", fg="red")
        self.lbl_excel.grid(row=1, column=0, sticky="w", pady=10)
        tk.Button(frame_arq, text="Selecionar Excel", command=self.load_excel).grid(row=1, column=1, padx=10)
        
        # Filtros
        frame_filt = tk.LabelFrame(root, text=" 2. Filtro Interativo (Screener) e Regras ", padx=10, pady=10)
        frame_filt.pack(fill="x", padx=20, pady=5)
        
        tk.Label(frame_filt, text="Escolha o Carro (Modelo):").grid(row=0, column=0, sticky="w")
        self.combo_modelo = ttk.Combobox(frame_filt, state="disabled", width=50)
        self.combo_modelo.grid(row=0, column=1, padx=10)
        
        tk.Label(frame_filt, text="Folha de Mapeamento:").grid(row=1, column=0, sticky="w", pady=15)
        tk.Button(frame_filt, text="Abrir / Configurar Regras", command=self.open_map).grid(row=1, column=1, sticky="w", padx=10, pady=15)
        
        # Gerar
        self.btn_run = tk.Button(root, text="GERAR PDF DO CARRO SELECIONADO", command=self.run_robot, bg="#4CAF50", fg="white", font=("Arial", 14, "bold"), height=2)
        self.btn_run.pack(fill="x", padx=20, pady=20)
        
        self.create_default_map()

    def create_default_map(self):
        if not os.path.exists(self.map_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Regras"
            ws.append(["TEXTO_NO_PDF_MOLDE", "TIPO_VALOR (Ex: REV01)", "COR_FUNDO_PDF", "COR_TEXTO_PDF"])
            ws.append(["1334 35", "REV01", "Branco", "Preto"])
            ws.append(["464,78", "REV02", "Vermelho Mit", "Branco"])
            for col in ['A', 'B', 'C', 'D']:
                ws.column_dimensions[col].width = 25
            wb.save(self.map_file)

    def load_pdf(self):
        p = filedialog.askopenfilename(filetypes=[("PDF", "*.pdf")])
        if p:
            self.pdf_molde = p
            self.lbl_pdf.config(text=f"PDF Molde: {os.path.basename(p)}", fg="green")

    def load_excel(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls *.csv")])
        if p:
            self.excel_dados = p
            self.lbl_excel.config(text="A ler milhares de linhas... aguarde.", fg="#ff9800")
            self.root.update()
            
            try:
                if p.endswith('.csv'):
                    self.df_base = pd.read_csv(p)
                else:
                    self.df_base = pd.read_excel(p, sheet_name="BASE_V_REDE")
                
                col_modelo = None
                for col in ["MODELO", "MODELO APLICADO", "CARRO"]:
                    if col in self.df_base.columns:
                        col_modelo = col
                        break
                
                if col_modelo:
                    modelos = sorted(self.df_base[col_modelo].dropna().unique().tolist())
                    self.combo_modelo['values'] = modelos
                    self.combo_modelo.config(state="readonly")
                    if modelos:
                        self.combo_modelo.current(0)
                    self.lbl_excel.config(text=f"Excel: {os.path.basename(p)}", fg="green")
                else:
                    messagebox.showerror("Erro", "Coluna 'MODELO' não encontrada na base de dados!")
                    self.lbl_excel.config(text="Erro ao ler Excel", fg="red")
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível ler a base de dados.\n{str(e)}")
                self.lbl_excel.config(text="Erro ao ler Excel", fg="red")

    def open_map(self):
        try:
            os.startfile(self.map_file)
        except:
            messagebox.showinfo("Aviso", f"Abra manualmente o ficheiro {self.map_file}")

    def get_color(self, name):
        name = str(name).strip().lower()
        if "branco" in name: return (1, 1, 1)
        if "preto" in name: return (0, 0, 0)
        if "vermelho" in name: return (0.83, 0.12, 0.15)
        return (1, 1, 1)

    def run_robot(self):
        if not self.pdf_molde or self.df_base is None:
            messagebox.showwarning("Atenção", "Selecione o PDF e o Excel primeiro!")
            return
            
        modelo_escolhido = self.combo_modelo.get()
        if not modelo_escolhido:
            messagebox.showwarning("Atenção", "Selecione um carro/modelo no filtro!")
            return
            
        try:
            wb_map = openpyxl.load_workbook(self.map_file)
            ws_map = wb_map.active
            regras = []
            for row in ws_map.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    regras.append({
                        "texto_pdf": str(row[0]).strip(),
                        "tipo_valor": str(row[1]).strip().upper(),
                        "fundo": row[2] if row[2] else "Branco",
                        "cor": row[3] if row[3] else "Preto"
                    })
                    
            if not regras:
                messagebox.showerror("Erro", "O mapeamento está vazio!")
                return
                
            col_modelo = "MODELO" if "MODELO" in self.df_base.columns else "MODELO APLICADO"
            df_filtrado = self.df_base[self.df_base[col_modelo] == modelo_escolhido]
            
            # Calcular matematicamente
            valores_calculados = {}
            if "Nº REVISÃO" in df_filtrado.columns and "VLR TT" in df_filtrado.columns:
                for rev in ["REV01", "REV02", "REV03", "REV04", "REV05", "REV06", "REV07", "REV08", "REV09", "REV10"]:
                    
                    # Convert to numeric in case of strings
                    soma = pd.to_numeric(df_filtrado[df_filtrado["Nº REVISÃO"] == rev]["VLR TT"], errors='coerce').sum()
                    
                    if soma > 0:
                        v_str = f"{soma:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                        valores_calculados[rev] = v_str
            
            doc = fitz.open(self.pdf_molde)
            subs = 0
            
            for page in doc:
                for regra in regras:
                    old = regra["texto_pdf"]
                    tipo = regra["tipo_valor"]
                    new = valores_calculados.get(tipo, "")
                    
                    if not new: continue
                    
                    instancias = page.search_for(old)
                    subs += len(instancias)
                    
                    bg_col = self.get_color(regra["fundo"])
                    fg_col = self.get_color(regra["cor"])
                    
                    for inst in instancias:
                        inst.x0 -= 1.5; inst.x1 += 1.5; inst.y0 -= 1; inst.y1 += 1
                        page.draw_rect(inst, color=bg_col, fill=bg_col)
                        page.insert_text(
                            (inst.x0 + 1, inst.y1 - inst.height * 0.25), 
                            new, 
                            fontsize=inst.height * 0.82, 
                            color=fg_col, fontname="helv"
                        )
                        
            save_path = filedialog.asksaveasfilename(defaultextension=".pdf", initialfile=f"Lamina_{modelo_escolhido.replace('/','_')}.pdf")
            if save_path:
                doc.save(save_path)
                messagebox.showinfo("Pronto!", f"A Versão Suprema finalizou com sucesso!\nForam calculados os valores baseados nos dados em bruto do {modelo_escolhido} e feitas {subs} atualizações automáticas na arte do PDF.")
                
        except Exception as e:
            messagebox.showerror("Erro Fatal", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    app = RoboSupremoApp(root)
    root.mainloop()
